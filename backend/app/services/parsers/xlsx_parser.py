"""XLSX Parser — 使用 openpyxl 解析 Excel 文档

支持:
- 逐 sheet 提取内容
- 小表整体转 Markdown table
- 大表(>50行)按行组切片，每组保留表头
- 保留 sheet_name 和 cell_range
- 提取合并单元格信息
- 提取批注
"""

import uuid
from pathlib import Path
from typing import Optional

from loguru import logger

from app.services.parsers.base import (
    BaseParser, ParserRegistry, UDRBlock, UnifiedDocument,
)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed; XLSX parsing will be limited")

# 大表行组切片参数
SMALL_TABLE_THRESHOLD = 50  # 50行以下为小表，整体输出
ROW_GROUP_SIZE = 30  # 大表每组30行


def _rows_to_markdown(headers: list[str], rows: list[list[str]]) -> str:
    """将 headers + rows 转为 Markdown table"""
    if not headers:
        return ""
    md_lines = ["| " + " | ".join(str(h) for h in headers) + " |"]
    md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        padded = list(row) + [""] * (len(headers) - len(row))
        md_lines.append("| " + " | ".join(str(c) for c in padded[:len(headers)]) + " |")
    return "\n".join(md_lines)


def _get_cell_range(row_start: int, row_end: int, col_count: int) -> str:
    """生成 Excel cell_range 字符串，如 A1:C10"""
    start_col = get_column_letter(1)
    end_col = get_column_letter(col_count) if col_count > 0 else "A"
    return f"{start_col}{row_start}:{end_col}{row_end}"


class XlsxParser(BaseParser):
    supported_mime_types = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ]
    supported_extensions = [".xlsx"]

    def parse(self, file_path: str, options: Optional[dict] = None) -> UnifiedDocument:
        if not HAS_OPENPYXL:
            return self._fallback_parse(file_path)

        doc_id = f"doc_{uuid.uuid4().hex[:12]}"
        path = Path(file_path)
        filename = path.name
        options = options or {}

        blocks: list[UDRBlock] = []

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Sheet heading
                blocks.append(UDRBlock(
                    block_id=f"{doc_id}_b{len(blocks):04d}",
                    type="heading",
                    text=f"Sheet: {sheet_name}",
                    level=1,
                    sheet_name=sheet_name,
                ))

                # 读取所有行
                rows_data: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    rows_data.append(cells)

                if not rows_data:
                    blocks.append(UDRBlock(
                        block_id=f"{doc_id}_b{len(blocks):04d}",
                        type="paragraph",
                        text="[Empty sheet]",
                        sheet_name=sheet_name,
                    ))
                    continue

                # 第一行作为表头
                headers = rows_data[0]
                data_rows = rows_data[1:]
                col_count = len(headers)

                if len(data_rows) <= SMALL_TABLE_THRESHOLD:
                    # 小表：整体输出
                    cell_range = _get_cell_range(1, len(data_rows) + 1, col_count)
                    md_text = _rows_to_markdown(headers, data_rows)

                    blocks.append(UDRBlock(
                        block_id=f"{doc_id}_b{len(blocks):04d}",
                        type="table",
                        text=md_text,
                        sheet_name=sheet_name,
                        cell_range=cell_range,
                        structured_data={
                            "headers": headers,
                            "rows": data_rows,
                            "total_rows": len(data_rows),
                        },
                    ))
                else:
                    # 大表：生成摘要 + 列名索引 + 行组切片
                    # 1. 摘要 block
                    cell_range_full = _get_cell_range(1, len(data_rows) + 1, col_count)
                    summary_text = (
                        f"Excel 表格「{sheet_name}」，共 {len(data_rows)} 行 × {col_count} 列。\n"
                        f"列名: {', '.join(str(h) for h in headers)}\n"
                        f"范围: {cell_range_full}"
                    )
                    blocks.append(UDRBlock(
                        block_id=f"{doc_id}_b{len(blocks):04d}",
                        type="paragraph",
                        text=summary_text,
                        sheet_name=sheet_name,
                        cell_range=cell_range_full,
                        metadata={"table_summary": True, "row_count": len(data_rows), "col_count": col_count},
                    ))

                    # 2. 列名索引 block
                    col_index_lines = [f"列 {i+1}: {h}" for i, h in enumerate(headers)]
                    blocks.append(UDRBlock(
                        block_id=f"{doc_id}_b{len(blocks):04d}",
                        type="paragraph",
                        text="列名索引:\n" + "\n".join(col_index_lines),
                        sheet_name=sheet_name,
                        metadata={"column_index": True, "columns": headers},
                    ))

                    # 3. 前5行预览 block
                    preview_rows = data_rows[:5]
                    preview_range = _get_cell_range(2, len(preview_rows) + 1, col_count)
                    blocks.append(UDRBlock(
                        block_id=f"{doc_id}_b{len(blocks):04d}",
                        type="table",
                        text=_rows_to_markdown(headers, preview_rows),
                        sheet_name=sheet_name,
                        cell_range=preview_range,
                        structured_data={"headers": headers, "rows": preview_rows},
                        metadata={"preview": True},
                    ))

                    # 4. 行组切片（每组 ROW_GROUP_SIZE 行，保留表头）
                    for group_start in range(0, len(data_rows), ROW_GROUP_SIZE):
                        group_rows = data_rows[group_start:group_start + ROW_GROUP_SIZE]
                        actual_row_start = group_start + 2  # +1 for header, +1 for 1-indexed
                        actual_row_end = actual_row_start + len(group_rows) - 1
                        group_cell_range = _get_cell_range(actual_row_start, actual_row_end, col_count)

                        blocks.append(UDRBlock(
                            block_id=f"{doc_id}_b{len(blocks):04d}",
                            type="table",
                            text=_rows_to_markdown(headers, group_rows),
                            sheet_name=sheet_name,
                            cell_range=group_cell_range,
                            structured_data={"headers": headers, "rows": group_rows},
                            metadata={"row_group": True, "row_start": group_start + 1, "row_end": group_start + len(group_rows)},
                        ))

                # 提取批注
                try:
                    ws_comments = getattr(ws, 'comments', None) or []
                    for comment in ws_comments:
                        comment_text = f"[批注] {comment.text}" if hasattr(comment, 'text') else str(comment)
                        blocks.append(UDRBlock(
                            block_id=f"{doc_id}_b{len(blocks):04d}",
                            type="annotation",
                            text=comment_text,
                            sheet_name=sheet_name,
                            metadata={"note_type": "comment"},
                        ))
                except Exception:
                    pass  # read_only mode may not support comments

            wb.close()

            return UnifiedDocument(
                document_id=doc_id,
                source={
                    "filename": filename,
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "source_uri": str(path.absolute()),
                },
                metadata={"title": filename, "sheet_count": len(wb.sheetnames)},
                blocks=blocks,
            )

        except Exception as e:
            logger.error(f"XLSX parsing failed: {e}")
            return self._fallback_parse(file_path)

    @staticmethod
    def _fallback_parse(file_path: str) -> UnifiedDocument:
        doc_id = f"doc_{uuid.uuid4().hex[:12]}"
        path = Path(file_path)
        return UnifiedDocument(
            document_id=doc_id,
            source={
                "filename": path.name,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "source_uri": str(path.absolute()),
            },
            metadata={"title": path.name, "warning": "openpyxl not available"},
            blocks=[
                UDRBlock(
                    block_id=f"{doc_id}_b0000",
                    type="paragraph",
                    text=f"[XLSX file: {path.name} — install openpyxl for full parsing]",
                ),
            ],
        )


# 注册
ParserRegistry.register(XlsxParser())
